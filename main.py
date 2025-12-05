import os
import glob
from datetime import date, datetime
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side

import xlrd
from xlrd import open_workbook, cellname
import xlwt
from xlwt import Workbook
import xlutils
from xlutils.copy import copy

from file_paths import invnday_path, po_us_vender_path, steel_codes_path, alum_codes_path, copper_codes_path, invnday_save_path #, metal_master_path, save_path, report_path
from chapter_98 import chapter_98_codes

today = date.today().isoformat() #.replace("-","")

list_of_invnday = glob.glob(invnday_path)
latest_invnday = max(list_of_invnday, key=os.path.getctime)

list_of_us_po_vendor = glob.glob(po_us_vender_path)
latest_us_po_vendor = max(list_of_us_po_vendor, key=os.path.getctime)

invnday_wb = open_workbook(latest_invnday)
invnday_sheet = invnday_wb.sheet_by_index(0)

final_invnday_wb = copy(invnday_wb)
final_invnday_ws = final_invnday_wb.get_sheet(0)

po_us_vendor_wb = openpyxl.load_workbook(latest_us_po_vendor)
po_us_vendor_ws = po_us_vendor_wb['Sheet1']

def harm_codes(file_path):  # Returns the contents of a text file as a list of strings from a file path.
    harm_codes = []
    with open(file_path) as f:
        file_contents = f.read()    
    for content in file_contents.split():
        harm_codes.append(content)
    return harm_codes 

def metal_declaration_check(row_index):
    for harm_code in harm_codes(alum_codes_path):
        value = invnday_sheet.cell(row_index, 10).value
        if value[:len(harm_code)] == harm_code:
            print(value, harm_code )
            print("declaration required")
            break

metal_declaration_check(42)

def acquire_chapter_98(harm_code):
    for key, value in chapter_98_codes.items():
        if re.match(key,harm_code):
            return value
        


def actual_po_date(po_row):
    po_date = po_us_vendor_ws.cell(row=po_row, column=2)
    char_list = list(str(po_date.value))
    char_list.insert(4, "-"), char_list.insert(7, "-")
    new_po_date = "".join(char_list)

    date_time_po_date = datetime.strptime(new_po_date, "%Y-%m-%d")
    date_time_today = datetime.strptime(today, "%Y-%m-%d")

    days_in_3_years = 1095

    if int(str(date_time_today - date_time_po_date).split()[0]) <= days_in_3_years:
        return True
    else:
        return False

def main():
    tracker = 0
    for row_index in range(1, invnday_sheet.nrows):
        invnday_sku = invnday_sheet.cell(row_index, 3).value
        start_of_chapter_98 = "9801"
        for po_row in range(1, po_us_vendor_ws.max_row + 1):
            po_sku = po_us_vendor_ws.cell(row=po_row, column=12).value
            if invnday_sku == po_sku:
                valid_for_chapter_98 = actual_po_date(po_row)
                if valid_for_chapter_98 == True:
                    
                    if invnday_sheet.cell(row_index, 10).value[:4] == start_of_chapter_98:
                        print("correct sku", row_index + 1)
                        tracker += 1
                    else:
                        final_invnday_ws.write(row_index, 10, acquire_chapter_98(invnday_sheet.cell(row_index, 10).value))
                        print("changed to", acquire_chapter_98(invnday_sheet.cell(row_index, 10).value), row_index + 1)
                        tracker += 1
                else:
                    if invnday_sheet.cell(row_index, 10).value[:4] == start_of_chapter_98:
                        print(f"Sku ({invnday_sheet.cell(row_index, 3).value}) has a PO purchase date that exceeds 3 years. Please change from chapter 98.", row_index + 1)
                        tracker += 1
                    else:
                        print("keep the same", row_index + 1)  # declaration function should go here.
                        tracker += 1
                break
            
        else:
            if invnday_sheet.cell(row_index, 10).value[:4] == start_of_chapter_98:
                print(f"Sku ({invnday_sheet.cell(row_index, 3).value}) has a PO purchase date that exceeds 3 years. Please change from chapter 98.", row_index + 1)
                tracker += 1
            
            else:
                print("keep the same", row_index + 1)  # declaration function should go here.
                tracker += 1
    print("tracker: ", tracker)

# main()


# final_invnday_save = latest_invnday.split("/")[-1]
# final_invnday_wb.save(f"{invnday_save_path}/final_{final_invnday_save}")